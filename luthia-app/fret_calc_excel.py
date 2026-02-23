import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import math

def calculate_fret_position(scale_length, fret_number):
    """Calculate fret position from nut using the formula"""
    return scale_length * (1 - 2**(-fret_number/12))

def create_fret_table(ws, scale_length, scale_name, num_frets, start_row):
    """Create a fret placement table in the worksheet"""
    
    # Header
    ws.merge_cells(f'A{start_row}:D{start_row}')
    header_cell = ws[f'A{start_row}']
    header_cell.value = f'{scale_name} - {num_frets} Frets'
    header_cell.font = Font(bold=True, size=12)
    header_cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_cell.font = Font(bold=True, size=12, color='FFFFFF')
    header_cell.alignment = Alignment(horizontal='center')
    
    # Column headers
    start_row += 1
    headers = ['Fret', 'Distance from Nut (mm)', 'Distance from Previous Fret (mm)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    prev_distance = 0
    for fret in range(num_frets + 1):
        row = start_row + fret + 1
        distance = calculate_fret_position(scale_length, fret) if fret > 0 else 0
        distance_from_prev = distance - prev_distance if fret > 0 else 0
        
        ws.cell(row=row, column=1).value = fret
        ws.cell(row=row, column=2).value = round(distance, 1)
        ws.cell(row=row, column=3).value = round(distance_from_prev, 1) if fret > 0 else '-'
        
        # Center align all cells
        for col in range(1, 4):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        prev_distance = distance
    
    return start_row + num_frets + 3  # Return next available row

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# Guitar scales
guitar_scales = [
    (648, '648mm (25.5" Fender)', [21, 22, 24]),
    (629, '629mm (24.75" Gibson)', [21, 22, 24])
]

# Bass scales
bass_scales = [
    (760, '760mm (30" Short Scale)', [24]),
    (810, '810mm (32" Medium Scale)', [24]),
    (860, '860mm (34" Standard Scale)', [24]),
    (889, '889mm (35" Extra-Long Scale)', [24])
]

# Create guitar sheets
for scale_length, scale_name, fret_counts in guitar_scales:
    ws = wb.create_sheet(title=f'{scale_length}mm Guitar')
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 35
    
    current_row = 1
    for num_frets in fret_counts:
        current_row = create_fret_table(ws, scale_length, scale_name, num_frets, current_row)

# Create bass sheets
for scale_length, scale_name, fret_counts in bass_scales:
    ws = wb.create_sheet(title=f'{scale_length}mm Bass')
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 35
    
    current_row = 1
    for num_frets in fret_counts:
        current_row = create_fret_table(ws, scale_length, scale_name, num_frets, current_row)

# Save the file
wb.save('fret_placement_tables.xlsx')
print("Excel file 'fret_placement_tables.xlsx' created successfully!")
