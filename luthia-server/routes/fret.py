"""Fret calculator routes.

Provides:
  GET  /fret                 — fret calculator page
  GET  /api/v1/fret/calculate   — compute fret positions (JSON)
  GET  /api/v1/fret/export      — download as .xlsx
"""

import io
import math

from flask import Blueprint, jsonify, render_template, request, send_file

fret_bp = Blueprint('fret', __name__)

# ---------------------------------------------------------------------------
# Preset scale lengths – guitar and bass standards
# Ordered as they appear in the UI preset buttons.
# ---------------------------------------------------------------------------
GUITAR_PRESETS = [
    {'label': '24.75″  Gibson',   'scale_mm': 628.65,  'instrument': 'guitar'},
    {'label': '25″  PRS',         'scale_mm': 635.0,   'instrument': 'guitar'},
    {'label': '25.5″  Fender',    'scale_mm': 647.7,   'instrument': 'guitar'},
    {'label': '26.5″  Baritone',  'scale_mm': 673.1,   'instrument': 'guitar'},
]

BASS_PRESETS = [
    {'label': '30″  Short',       'scale_mm': 762.0,   'instrument': 'bass'},
    {'label': '32″  Medium',      'scale_mm': 812.8,   'instrument': 'bass'},
    {'label': '34″  Standard',    'scale_mm': 863.6,   'instrument': 'bass'},
    {'label': '35″  Extra Long',  'scale_mm': 889.0,   'instrument': 'bass'},
]

# ---------------------------------------------------------------------------
# Page route
# ---------------------------------------------------------------------------

@fret_bp.route('/fret')
def fret_index():
    """Render the fret calculator page."""
    return render_template(
        'fret.html',
        guitar_presets=GUITAR_PRESETS,
        bass_presets=BASS_PRESETS,
        active_nav='fret',
        breadcrumb=[('Fret Calculator', None)],
        page_title='Luthia · Fret Calculator',
    )


# ---------------------------------------------------------------------------
# Calculation API  (also used server-side for the export)
# ---------------------------------------------------------------------------

@fret_bp.route('/api/v1/fret/calculate', endpoint='api_fret_calculate')
def api_fret_calculate():
    """Return fret position data as JSON.

    Query params:
      scale_mm   — scale length in mm (required, float)
      num_frets  — number of frets (default 24, int)
    """
    scale_mm  = request.args.get('scale_mm',  type=float)
    num_frets = request.args.get('num_frets', 24, type=int)

    if not scale_mm or scale_mm <= 0:
        return jsonify({'ok': False, 'error': 'scale_mm must be a positive number.'}), 400
    if not (1 <= num_frets <= 36):
        return jsonify({'ok': False, 'error': 'num_frets must be between 1 and 36.'}), 400

    frets = _compute_frets(scale_mm, num_frets)
    return jsonify({'ok': True, 'scale_mm': scale_mm, 'num_frets': num_frets, 'frets': frets})


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

@fret_bp.route('/api/v1/fret/export', endpoint='api_fret_export')
def api_fret_export():
    """Generate and return a .xlsx fret table for download.

    Query params: same as /fret/calculate, plus optional label (string).
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return 'openpyxl is not installed.', 500

    scale_mm  = request.args.get('scale_mm',  type=float)
    num_frets = request.args.get('num_frets', 24, type=int)
    label     = request.args.get('label', '').strip() or f'{scale_mm:.1f}mm'

    if not scale_mm or scale_mm <= 0:
        return 'Invalid scale_mm', 400

    frets = _compute_frets(scale_mm, num_frets)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fret Positions'

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 18

    # Title header
    ws.merge_cells('A1:D1')
    h = ws['A1']
    h.value = f'{label} — {num_frets} Frets'
    h.font  = Font(bold=True, size=12, color='FFFFFF')
    h.fill  = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    h.alignment = Alignment(horizontal='center')

    # Column headers
    col_headers = ['Fret', 'Distance from Nut (mm)', 'Spacing from Prev. Fret (mm)', 'Distance from Nut (in)']
    for c, ch in enumerate(col_headers, 1):
        cell = ws.cell(row=2, column=c)
        cell.value = ch
        cell.font  = Font(bold=True, size=10)
        cell.fill  = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, f in enumerate(frets, 3):
        ws.cell(row=row_idx, column=1).value = f['fret']
        ws.cell(row=row_idx, column=2).value = round(f['from_nut_mm'],   2)
        ws.cell(row=row_idx, column=3).value = '' if f['spacing_mm'] is None else round(f['spacing_mm'], 2)
        ws.cell(row=row_idx, column=4).value = round(f['from_nut_in'],   4)
        for c in range(1, 5):
            ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal='center')
        # Alternate row shading
        if row_idx % 2 == 0:
            for c in range(1, 5):
                ws.cell(row=row_idx, column=c).fill = PatternFill(
                    start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_label = label.replace('/', '-').replace(' ', '_')
    filename   = f'fret_table_{safe_label}_{num_frets}frets.xlsx'

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ---------------------------------------------------------------------------
# Pure-Python calculation (no DB, no Flask)
# ---------------------------------------------------------------------------

def _compute_frets(scale_mm: float, num_frets: int) -> list[dict]:
    """Return a list of fret dicts, one entry per fret (0 = nut).

    Each dict contains:
      fret         — fret number (0 = nut)
      from_nut_mm  — distance from nut in mm
      from_nut_in  — distance from nut in inches
      spacing_mm   — distance from previous fret (None for fret 0)
      spacing_in   — spacing in inches (None for fret 0)
    """
    result = []
    prev_mm = 0.0

    for n in range(num_frets + 1):
        if n == 0:
            mm = 0.0
        else:
            mm = scale_mm * (1.0 - math.pow(2.0, -n / 12.0))

        spacing_mm = (mm - prev_mm) if n > 0 else None

        result.append({
            'fret':       n,
            'from_nut_mm': round(mm, 3),
            'from_nut_in': round(mm / 25.4, 4),
            'spacing_mm':  round(spacing_mm, 3) if spacing_mm is not None else None,
            'spacing_in':  round(spacing_mm / 25.4, 4) if spacing_mm is not None else None,
        })
        prev_mm = mm

    return result